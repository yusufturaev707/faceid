from io import BytesIO
from django.contrib import admin
from django.db import models
from django.contrib.postgres.fields import ArrayField
from django import forms
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect
from django.urls import path
import pandas as pd
from django.contrib import messages
import openpyxl
from import_export.admin import ExportActionModelAdmin, ImportExportModelAdmin

from unfold.admin import ModelAdmin
from unfold.contrib.forms.widgets import WysiwygWidget, ArrayWidget, UnfoldAdminTextInputWidget, UnfoldAdminSelectWidget
from unfold.decorators import action, display
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from unfold.enums import ActionVariant
from unfold.paginator import InfinitePaginator

from access_control.models import Staff, Supervisor, NormalUserLog, EventSupervisor
from core.utils import get_image_from_personal_info, get_personal_data
from region.models import Region
from users.models import User


class ExcelImportForm(forms.Form):
    excel_file = forms.FileField(
        label='📤 Import uchun Excel fayl tanlang',
        help_text='Faqat **.xlsx** yoki **.xls** formatidagi fayllarni yuklang.',
        widget=forms.FileInput(attrs={
            # Tailwind/Zamonaviy ko'rinish uchun klasslar
            'class': 'block w-full text-sm text-gray-500 ' # Asosiy stili
                     'file:mr-4 file:py-2 file:px-4 ' # Fayl tanlash tugmasini stilini beradi
                     'file:rounded-full file:border-0 '
                     'file:text-sm file:font-semibold '
                     'file:bg-indigo-50 file:text-indigo-700 ' # Tugma rangi
                     'hover:file:bg-indigo-100', # Hover effekti
            'accept': '.xlsx,.xls',
        })
    )


@admin.register(Supervisor)
class SupervisorAdmin(ModelAdmin, ExportActionModelAdmin):
    list_display = ['id', 'full_name', 'imei', 'gender', 'region']
    list_filter = ['status', 'gender']
    readonly_fields = ['id', 'created_at', 'updated_at', 'image_tag']
    search_fields = ['last_name', 'first_name', 'middle_name', 'imei']
    list_display_links = ['id', 'full_name', 'region', 'image_tag', 'imei']
    list_per_page = 30

    actions_list = ['import_changelist_action']
    actions_detail = ["get_image_gsp_action"]

    def get_queryset(self, request):
        qs = super().get_queryset(request).order_by('-id')
        user = User.objects.get(id=request.user.id)

        if user.is_admin or user.is_central:
            return qs

        user_region = getattr(request.user, "region", None)
        if user_region:
            return qs.filter(region=user_region)
        return qs.none()

    fieldsets = (
        ('Rasm', {
            'fields': ('image_tag',)
        }),
        ('Shaxsiy ma\'lumotlar', {
            'fields': ('last_name', 'first_name', 'middle_name', 'gender', 'ps_ser', 'ps_num', 'imei',)
        }),
        ('Viloyat', {
            'fields': ('region', )
        }),
        ('Holat', {
            'fields': ('status', )
        }),
        ('Time', {
            'fields': ('created_at', 'updated_at')
        }),
    )

    @display(description='F.I.O')
    def full_name(self, obj):
        return f"{obj.last_name} {obj.first_name} {obj.middle_name or ''}".strip().upper()

    @display(description='PASPORT')
    def ps_data(self, obj):
        return f"{obj.ps_ser} {obj.ps_num}"

    @display(description='STATUS')
    def status(self, obj):
        return obj.status

    @action(
        description="Rasmni yangilash",
        permissions=["get_image_gsp_action"],
    )
    def get_image_gsp_action(self, request, object_id):
        try:
            object_id = int(object_id)
        except (TypeError, ValueError) as e:
            raise Http404 from e

        try:
            supervisor = Supervisor.objects.get(id=int(object_id))
            ps_num = supervisor.ps_num
            nol_count = 7 - len(ps_num)
            ps_num = "0" * nol_count + ps_num
            data: dict = get_personal_data(imi=supervisor.imei, ps=f"{supervisor.ps_ser}{ps_num}")
            if data['status'] == 1:
                supervisor.img_b64 = data['photo']
                supervisor.last_name = data['sname']
                supervisor.first_name = data['fname']
                supervisor.middle_name = data['mname']
                supervisor.gender = 'M' if data['sex'] == 1 else 'F'
                supervisor.save()
                messages.success(request,f"{data['message']}")
            else:
                messages.error(request, f"{data['message']}")
        except Supervisor.DoesNotExist as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, str(e))
        finally:
            return redirect(request.headers["referer"])

    def has_get_image_gsp_action_permission(self, request, object_id):
        return request.user.is_admin or request.user.is_central or request.user.is_delegate

    @action(description=_("Shablon"), icon="download",  variant=ActionVariant.PRIMARY)
    def template_changelist_action(self, request):
        messages.success(
            request, _("Shablon yuklab olindi!")
        )
        return redirect(reverse_lazy("admin:person_download_template"))


    @action(description=_("Import"), icon="upload", variant=ActionVariant.PRIMARY,  permissions=["import_changelist_action"],)
    def import_changelist_action(self, request):
        return redirect(reverse_lazy("admin:access_control_supervisor_import_excel"))


    def has_import_changelist_action_permission(self, request):
        return request.user.is_admin or request.user.is_central or request.user.is_delegate

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("import-excel/", self.admin_site.admin_view(self.import_excel_view),
                 name="access_control_supervisor_import_excel"),
            path('download-template/', self.download_template, name='person_download_template'),
        ]
        return custom_urls + urls

    def import_excel_view(self, request):
        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']

                user = User.objects.get(id=request.user.id)

                try:
                    df = pd.read_excel(excel_file)

                    success_count = 0
                    error_count = 0
                    errors = []

                    for index, row in df.iterrows():
                        try:
                            person_data = {
                                'first_name': str(row.get('Ism', '')).strip(),
                                'last_name': str(row.get('Familiya', '')).strip(),
                                'middle_name': str(row.get('Otasining ismi', '')).strip() if pd.notna(
                                    row.get('Otasining ismi')) else '',
                                'gender': str(row.get('Jinsi', 'M')).strip().upper(),
                                'ps_ser': str(row.get('Pasport seriya', '')).strip(),
                                'ps_num': str(row.get('Pasport raqam', '')).strip(),
                                'imei': str(row.get('PINFL', '')).strip() if pd.notna(row.get('PINFL')) else '',
                                'phone': str(row.get('Telefon', '')).strip(),
                                'region_id': user.region.id,
                            }

                            # Majburiy maydonlarni tekshirish
                            if not person_data['first_name'] or not person_data['last_name']:
                                errors.append(f"Qator {index + 2}: Ism va Familiya bo'sh bo'lishi mumkin emas")
                                error_count += 1
                                continue

                            existing_person = Supervisor.objects.filter(imei=person_data['imei']).first()

                            if existing_person:
                                # Mavjud bo'lsa, yangilash
                                for key, value in person_data.items():
                                    setattr(existing_person, key, value)
                                existing_person.save()
                            else:
                                # Yangi yaratish
                                existing_person = Supervisor.objects.create(**person_data)

                            nol_count = 7 - len(person_data['ps_num'])
                            ps_num = "0" * nol_count + person_data['ps_num']

                            data: dict = get_personal_data(imi=existing_person.imei, ps=f"{existing_person.ps_ser}{ps_num}")
                            if data['status'] == 1:
                                existing_person.img_b64 = data['photo']
                                existing_person.last_name = data['sname']
                                existing_person.first_name = data['fname']
                                existing_person.middle_name = data['mname']
                                existing_person.gender = 'M' if data['sex'] == 1 else 'F'
                                existing_person.save()
                            else:
                                messages.error(request, f"{data['message']}")
                            success_count += 1
                        except Exception as e:
                            error_count += 1
                            errors.append(f"Qator {index + 2}: {str(e)}")

                    # Xabar ko'rsatish
                    if success_count > 0:
                        messages.success(request, f"✅ Muvaffaqiyatli yuklandi: {success_count} ta shaxs")

                    if error_count > 0:
                        error_message = f"❌ Xatoliklar: {error_count} ta\n" + "\n".join(errors[:10])
                        if len(errors) > 10:
                            error_message += f"\n... va yana {len(errors) - 10} ta xatolik"
                        messages.error(request, error_message)

                    return redirect('..')

                except Exception as e:
                    messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'title': 'Excel fayldan shaxslarni yuklash',
            'site_header': 'Person Import',
            'has_view_permission': True,
        }
        return render(request, "admin/import_excel.html", context)


    def download_template(self, request):
        # Excel fayl yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Person Import"

        # Sarlavhalar
        headers = [
            'Familiya', 'Ism', 'Otasining ismi', 'Jinsi', 'Pasport seriya', 'Pasport raqam', 'PINFL','Telefon'
        ]
        ws.append(headers)

        # Namuna ma'lumot
        sample_data = [
            'Oqboyev', 'Qoraboy', 'Sariq o\'g\'li', 'M', 'AA', '1234567', '12345678901234', '+998901234567'
        ]
        ws.append(sample_data)

        # Sarlavhalarni qalin qilish
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Ustunlar kengligini sozlash
        column_widths = [15, 15, 20, 15, 8, 15, 15, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Faylni saqlash
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=person_shablon.xlsx'

        return response

    compressed_fields = True
    list_fullwidth = False
    list_filter_sheet = True
    list_horizontal_scrollbar_bottom = True
    paginator = InfinitePaginator
    show_full_result_count = True

    formfield_overrides = {
        models.ForeignKey: {
            "widget": UnfoldAdminSelectWidget,
        },
    }


@admin.register(EventSupervisor)
class EventSupervisorAdmin(ModelAdmin, ExportActionModelAdmin):
    list_display = ['id', 'supervisor', 'exam', 'zone', 'category_name', 'test_date', 'sm', 'group_n', 'is_participated']
    list_filter = ['is_participated', ]
    readonly_fields = ['id', 'created_at', 'updated_at', ]
    search_fields = ['supervisor__last_name', 'supervisor__first_name', 'supervisor__middle_name', 'supervisor__imei']
    list_display_links = ['id', 'supervisor', 'exam', 'zone', 'category_name']
    list_per_page = 30

    def get_queryset(self, request):
        qs = super().get_queryset(request).order_by('-id')
        user = User.objects.get(id=request.user.id)

        if user.is_admin or user.is_central:
            return qs

        user_region = getattr(request.user, "region", None)
        if user_region:
            return qs.filter(supervisor__region=user_region)
        return qs.none()

    # fieldsets = (
    #     ('Shaxsiy ma\'lumotlar', {
    #         'fields': ('supervisor__last_name', 'supervisor__first_name', 'supervisor__middle_name', 'supervisor__gender', 'supervisor__ps_ser', 'supervisor__ps_num', 'supervisor__imei',)
    #     }),
    #     ('Viloyat va bino', {
    #         'fields': ('supervisor__region', 'zone', )
    #     }),
    #     ('Holat', {
    #         'fields': ('supervisor__status', 'is_participated')
    #     }),
    #     ('Time', {
    #         'fields': ('created_at', 'updated_at')
    #     }),
    # )

    compressed_fields = True
    list_fullwidth = False
    list_filter_sheet = True
    list_horizontal_scrollbar_bottom = True
    paginator = InfinitePaginator
    show_full_result_count = True

    formfield_overrides = {
        models.ForeignKey: {
            "widget": UnfoldAdminSelectWidget,
        },
    }


@admin.register(NormalUserLog)
class NormalUserLogAdmin(ModelAdmin):
    list_display = ['id', 'normal_user_id', 'last_name', 'first_name', 'middle_name', 'normal_user_type', 'get_region', 'zone', 'employee_no', 'ip_address', 'mac_address', 'direction', 'status', 'pass_time',]
    list_filter = ['door', 'direction', 'status']
    readonly_fields = ['id', 'created_at', 'updated_at', 'image_tag', 'get_region']
    search_fields = ['employee_no', 'ip_address', 'mac_address', 'direction', 'status']
    list_display_links = ['id', 'ip_address', 'mac_address', 'direction', 'status', 'employee_no']
    list_per_page = 50


    compressed_fields = True
    list_fullwidth = False
    list_filter_sheet = True
    list_horizontal_scrollbar_bottom = True
    paginator = InfinitePaginator
    show_full_result_count = False

    fieldsets = (
        (None, {
            'fields': ('employee_no', 'normal_user_type', 'last_name', 'first_name', 'middle_name', 'normal_user_id',)
        }),
        ('Hudud', {
            'fields': ('get_region', 'zone',)
        }),
        ('Turniket', {
            'fields': ('mac_address', 'ip_address', 'direction', 'status', 'pass_time',)
        }),
        ('Time', {
            'fields': ('created_at', 'updated_at')
        }),
        ('Rasm', {
            'fields': ('image_tag',)
        }),
    )

    @display(description='Bino')
    def get_zone(self, obj):
        return f"{obj.zone.name}" if obj.zone else ''

    formfield_overrides = {
        ArrayField: {
            "widget": ArrayWidget,
        },
        models.ForeignKey: {
            "widget": UnfoldAdminSelectWidget,
        },
        models.TimeField: {
            "widget": UnfoldAdminTextInputWidget(attrs={"type": "time"}),
        },
    }


@admin.register(Staff)
class StaffAdmin(ModelAdmin, ExportActionModelAdmin):
    list_display = ['id', 'image_tag', 'last_name', 'first_name', 'middle_name', 'imei', 'region', 'gender', 'status']
    list_filter = ['status', 'gender', 'region']
    readonly_fields = ['id', 'created_at', 'updated_at', 'image_tag']
    search_fields = ['last_name', 'first_name', 'middle_name', 'imei']
    list_display_links = ['id', 'last_name', 'first_name', 'middle_name', 'imei', 'phone']
    list_per_page = 30

    actions_list = ['import_changelist_action']
    actions_detail = ["get_image_gsp_action"]

    compressed_fields = True
    list_fullwidth = False
    list_filter_sheet = True
    list_horizontal_scrollbar_bottom = True
    paginator = InfinitePaginator
    show_full_result_count = False


    def get_queryset(self, request):
        qs = super().get_queryset(request).order_by('-id')
        user = User.objects.get(id=request.user.id)

        if user.is_admin or user.is_central:
            return qs

        user_region = getattr(request.user, "region", None)
        if user_region:
            return qs.filter(region=user_region)
        return qs.none()

    formfield_overrides = {
        models.TimeField: {
            "widget": UnfoldAdminTextInputWidget(attrs={"type": "time"}),
        },
    }

    fieldsets = (
        ('Shaxsiy ma\'lumotlar', {
            'fields': ('first_name', 'last_name', 'middle_name', 'phone', 'region', 'gender')
        }),
        ('Pasport ma\'lumotlari', {
            'fields': ('ps_ser', 'ps_num', 'imei')
        }),
        ('Holat', {
            'fields': ('status', )
        }),
        ('Rasm', {
            'fields': ('image_tag',)
        }),
    )

    @display(description='F.I.O')
    def full_name(self, obj):
        return f"{obj.last_name} {obj.first_name} {obj.middle_name or ''}".strip().upper()

    @display(description='VILOYAT')
    def region(self, obj):
        return f"{obj.name}".capitalize()

    @display(description='PASPORT')
    def ps_data(self, obj):
        return f"{obj.ps_ser} {obj.ps_num}"

    @display(description='TELEFON')
    def phone(self, obj):
        return f"{obj.phone}".strip()

    @display(description='STATUS')
    def status(self, obj):
        return obj.status

    @action(
        description="Rasmni yangilash",
        permissions=["get_image_gsp_action"],
    )
    def get_image_gsp_action(self, request, object_id):
        try:
            object_id = int(object_id)
        except (TypeError, ValueError) as e:
            raise Http404 from e

        try:
            staff = Staff.objects.get(id=int(object_id))
            ps_num = staff.ps_num
            nol_count = 7 - len(ps_num)
            ps_num = "0" * nol_count + ps_num
            data: dict = get_personal_data(imi=staff.imei, ps=f"{staff.ps_ser}{ps_num}")
            if data['status'] == 1:
                staff.img_b64 = data['photo']
                staff.last_name = data['sname']
                staff.first_name = data['fname']
                staff.middle_name = data['mname']
                staff.gender = 'M' if data['sex'] == 1 else 'F'
                staff.save()
                messages.success(request, f"{data['message']}")
            else:
                messages.error(request, f"{data['message']}")
        except Staff.DoesNotExist as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, str(e))
        finally:
            return redirect(request.headers["referer"])

    def has_get_image_gsp_action_permission(self, request, object_id):
        return request.user.is_admin or request.user.is_central or request.user.is_delegate

    @action(description=_("Shablon"), icon="download", variant=ActionVariant.PRIMARY)
    def template_changelist_action(self, request):
        messages.success(
            request, _("Shablon yuklab olindi!")
        )
        return redirect(reverse_lazy("admin:person_download_template"))

    @action(description=_("Import"), icon="upload", variant=ActionVariant.PRIMARY,
            permissions=["import_changelist_action"], )
    def import_changelist_action(self, request):
        return redirect(reverse_lazy("admin:access_control_staff_import_excel"))

    def has_import_changelist_action_permission(self, request):
        return request.user.is_admin or request.user.is_central or request.user.is_delegate

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("import-excel/", self.admin_site.admin_view(self.import_excel_view),
                 name="access_control_staff_import_excel"),
            path('download-template/', self.download_template, name='person_download_template'),
        ]
        return custom_urls + urls

    def import_excel_view(self, request):
        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']

                try:
                    df = pd.read_excel(excel_file)

                    success_count = 0
                    error_count = 0
                    errors = []

                    for index, row in df.iterrows():
                        try:
                            region = Region.objects.get(number=int(str(row.get('PCode', 14)).strip()))
                            person_data = {
                                'first_name': str(row.get('Ism', '')).strip(),
                                'last_name': str(row.get('Familiya', '')).strip(),
                                'middle_name': str(row.get('Otasining ismi', '')).strip() if pd.notna(
                                    row.get('Otasining ismi')) else '',
                                'gender': str(row.get('Jinsi', 'M')).strip().upper(),
                                'ps_ser': str(row.get('Pasport seriya', '')).strip(),
                                'ps_num': str(row.get('Pasport raqam', '')).strip(),
                                'imei': str(row.get('PINFL', '')).strip() if pd.notna(row.get('PINFL')) else '',
                                'phone': str(row.get('Telefon', '')).strip(),
                                'region_id': region.id,
                            }

                            # Majburiy maydonlarni tekshirish
                            if not person_data['first_name'] or not person_data['last_name']:
                                errors.append(f"Qator {index + 2}: Ism va Familiya bo'sh bo'lishi mumkin emas")
                                error_count += 1
                                continue

                            existing_person = Staff.objects.filter(imei=person_data['imei']).first()
                            if existing_person:
                                # Mavjud bo'lsa, yangilash
                                for key, value in person_data.items():
                                    setattr(existing_person, key, value)
                                existing_person.save()
                            else:
                                # Yangi yaratish
                                existing_person = Staff.objects.create(**person_data)

                            nol_count = 7 - len(person_data['ps_num'])
                            ps_num = "0" * nol_count + person_data['ps_num']

                            try:
                                data: dict = get_personal_data(imi=existing_person.imei,
                                                               ps=f"{existing_person.ps_ser}{ps_num}")
                            except Exception as e:
                                print(f"Asosiy xatolik: {e}")
                            if data['status'] == 1:
                                existing_person.img_b64 = data['photo']
                                existing_person.last_name = data['sname']
                                existing_person.first_name = data['fname']
                                existing_person.middle_name = data['mname']
                                existing_person.gender = 'M' if data['sex'] == 1 else 'F'
                                existing_person.save()
                            else:
                                messages.error(request, f"{data['message']}")

                            success_count += 1
                        except Exception as e:
                            error_count += 1
                            errors.append(f"Qator {index + 2}: {str(e)}")

                    # Xabar ko'rsatish
                    if success_count > 0:
                        messages.success(request, f"✅ Muvaffaqiyatli yuklandi: {success_count} ta shaxs")

                    if error_count > 0:
                        error_message = f"❌ Xatoliklar: {error_count} ta\n" + "\n".join(errors[:10])
                        if len(errors) > 10:
                            error_message += f"\n... va yana {len(errors) - 10} ta xatolik"
                        messages.error(request, error_message)

                    return redirect('..')

                except Exception as e:
                    messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'title': 'Excel fayldan shaxslarni yuklash',
            'site_header': 'Person Import',
            'has_view_permission': True,
        }
        return render(request, "admin/import_excel.html", context)

    def download_template(self, request):
        # Excel fayl yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Person Import"

        # Sarlavhalar
        headers = [
            'Familiya', 'Ism', 'Otasining ismi', 'Jinsi', 'Pasport seriya', 'Pasport raqam', 'PINFL', 'Telefon', 'PCode'
        ]
        ws.append(headers)

        # Namuna ma'lumot
        sample_data = [
            'Oqboyev', 'Qoraboy', 'Sariq o\'g\'li', 'M', 'AA', '1234567', '12345678901234', '+998901234567', 'Toshkent shahri'
        ]
        ws.append(sample_data)

        # Sarlavhalarni qalin qilish
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Ustunlar kengligini sozlash
        column_widths = [15, 15, 20, 15, 8, 15, 15, 18, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Faylni saqlash
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=person_shablon.xlsx'

        return response