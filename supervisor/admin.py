import time
from io import BytesIO
from django.contrib import admin
from django.db import models
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect
from django.urls import path
import pandas as pd
from django.contrib import messages
import openpyxl
from import_export.admin import ExportActionModelAdmin

from region.models import Region, SwingBarrier
from supervisor.utils import is_check_healthy, add_supervisor_to_swing_barr, upload_single_supervisor_face_image
from supervisor.forms import ExcelImportForm

from unfold.admin import ModelAdmin
from unfold.contrib.forms.widgets import UnfoldAdminTextInputWidget, UnfoldAdminSelectWidget
from unfold.decorators import action, display
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from unfold.enums import ActionVariant
from unfold.paginator import InfinitePaginator

from supervisor.models import Supervisor, EventSupervisor
from core.utils import get_personal_data
from users.models import User



@admin.register(Supervisor)
class SupervisorAdmin(ModelAdmin, ExportActionModelAdmin):
    list_display = ['id', 'full_name', 'imei', 'gender', 'region', 'role', 'status']
    list_filter = ['status', 'gender', 'role']
    readonly_fields = ['id', 'created_at', 'updated_at', 'image_tag']
    search_fields = ['last_name', 'first_name', 'middle_name', 'imei']
    list_display_links = ['id', 'full_name', 'region', 'image_tag', 'imei']
    list_per_page = 30

    actions_list = ['import_changelist_action', 'send_data_action']
    actions_detail = ["get_image_gsp_action"]

    def get_queryset(self, request):
        qs = super().get_queryset(request).order_by('-id')
        user = User.objects.get(id=request.user.id)

        if user.is_admin or user.is_central:
            return qs

        user_region = getattr(request.user, "region", None)
        if user_region:
            return qs.filter(region=user_region)
        return qs.none()

    fieldsets = (
        ('Rasm', {
            'fields': ('image_tag',)
        }),
        ('Shaxsiy ma\'lumotlar', {
            'fields': ('last_name', 'first_name', 'middle_name', 'gender', 'ps_ser', 'ps_num', 'imei',)
        }),
        (None , {
            'fields': ('region', 'role', 'status')
        }),
        ('Time', {
            'fields': ('created_at', 'updated_at')
        }),
    )

    @display(description='F.I.O')
    def full_name(self, obj):
        return f"{obj.last_name} {obj.first_name} {obj.middle_name or ''}".strip().upper()

    @display(description='PASPORT')
    def ps_data(self, obj):
        return f"{obj.ps_ser} {obj.ps_num}"

    @display(description='STATUS')
    def status(self, obj):
        return obj.status

    @action(
        description="Rasmni yangilash",
        permissions=["get_image_gsp_action"],
    )
    def get_image_gsp_action(self, request, object_id):
        try:
            object_id = int(object_id)
        except (TypeError, ValueError) as e:
            raise Http404 from e

        try:
            supervisor = Supervisor.objects.get(id=int(object_id))
            ps_num = supervisor.ps_num
            nol_count = 7 - len(ps_num)
            ps_num = "0" * nol_count + ps_num
            data: dict = get_personal_data(imi=supervisor.imei, ps=f"{supervisor.ps_ser}{ps_num}")
            if data['status'] == 1:
                supervisor.img_b64 = data['photo']
                supervisor.last_name = data['sname']
                supervisor.first_name = data['fname']
                supervisor.middle_name = data['mname']
                supervisor.gender = 'M' if data['sex'] == 1 else 'F'
                supervisor.save()
                messages.success(request,f"{data['message']}")
            else:
                messages.error(request, f"{data['message']}")
        except Supervisor.DoesNotExist as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, str(e))
        finally:
            return redirect(request.headers["referer"])

    def has_get_image_gsp_action_permission(self, request, object_id):
        return request.user.is_admin or request.user.is_central or request.user.is_delegate

    @action(description=_("Face ID"), icon="ar_on_you", variant=ActionVariant.SUCCESS,
            permissions=["send_data_action"], )
    def send_data_action(self, request):
        try:
            staff_queryset = Supervisor.objects.filter(role='staff').order_by('region', 'id')
            error_message = str()

            for staff in staff_queryset:
                region = staff.region
                imei = staff.imei
                img64 = staff.img_b64

                sb_queryset = SwingBarrier.objects.filter(zone__region=region).order_by('id')

                for sb in sb_queryset:
                    sb_ip = sb.ip_address
                    sb_username = sb.username
                    sb_password = sb.password
                    sb_mac_address = sb.mac_address
                    status = is_check_healthy(ip_address=sb_ip, mac_address=sb_mac_address, username=sb_username, password=sb_password)
                    if not status:
                        error_message += f"{staff.fio} - {region.name} - {sb_ip} - holati faol emas!\n"
                        continue
                    try:
                        is_user_added = add_supervisor_to_swing_barr(sb_ip, sb_username, sb_password, staff)

                        if is_user_added:
                            time.sleep(0.2)

                            img_data = {"fpid": imei, "img64": img64}

                            is_image_uploaded = upload_single_supervisor_face_image(
                                user_data=img_data,
                                ip_address=sb_ip,
                                username=sb_username,
                                password=sb_password
                            )

                            if is_image_uploaded:
                                print(f"✓ User {imei} va rasmi yuklandi")
                            else:
                                error_message += f"{staff.fio} - {region.name} - {sb_ip} - Rasm qo'shilmadi!\n"
                        else:
                            error_message += f"{staff.fio} - {region.name} - {sb_ip} - qo'shilmadi!\n"
                    except Exception as e:
                        error_message += f"{staff.fio} - {region.name} - {sb_ip} - {str(e)}\n"

        except Exception as e:
            messages.error(request, str(e))
        finally:
            return redirect(request.headers["referer"])

    def has_send_data_action_permission(self, request):
        return request.user.is_admin or request.user.is_central

    @action(description=_("Shablon"), icon="download",  variant=ActionVariant.PRIMARY)
    def template_changelist_action(self, request):
        messages.success(
            request, _("Shablon yuklab olindi!")
        )
        return redirect(reverse_lazy("admin:person_download_template"))


    @action(description=_("Import"), icon="upload", variant=ActionVariant.PRIMARY,  permissions=["import_changelist_action"],)
    def import_changelist_action(self, request):
        return redirect(reverse_lazy("admin:supervisor_supervisor_import_excel"))


    def has_import_changelist_action_permission(self, request):
        return request.user.is_admin or request.user.is_central or request.user.is_delegate

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("import-excel/", self.admin_site.admin_view(self.import_excel_view),
                 name="supervisor_supervisor_import_excel"),
            path('download-template/', self.download_template, name='person_download_template'),
        ]
        return custom_urls + urls

    def import_excel_view(self, request):
        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']

                user = User.objects.get(id=request.user.id)

                try:
                    df = pd.read_excel(excel_file)

                    success_count = 0
                    error_count = 0
                    errors = []

                    for index, row in df.iterrows():
                        try:
                            region = Region.objects.get(number=int(row.get('Pcode')))
                            person_data = {
                                'first_name': str(row.get('Ism', '')).strip(),
                                'last_name': str(row.get('Familiya', '')).strip(),
                                'middle_name': str(row.get('Otasining ismi', '')).strip() if pd.notna(
                                    row.get('Otasining ismi')) else '',
                                'gender': str(row.get('Jinsi', 'M')).strip().upper(),
                                'ps_ser': str(row.get('Pasport seriya', '')).strip(),
                                'ps_num': str(row.get('Pasport raqam', '')).strip(),
                                'imei': str(row.get('PINFL', '')).strip() if pd.notna(row.get('PINFL')) else '',
                                'region_id':region.id if user.is_admin or user.is_central else user.region.id,
                                'role': str(row.get('Role', 'supervisor')).strip() if user.is_admin or user.is_central else 'supervisor',
                            }

                            # Majburiy maydonlarni tekshirish
                            if not person_data['first_name'] or not person_data['last_name']:
                                errors.append(f"Qator {index + 2}: Ism va Familiya bo'sh bo'lishi mumkin emas")
                                error_count += 1
                                continue

                            existing_person = Supervisor.objects.filter(imei=person_data['imei']).first()

                            if existing_person:
                                # Mavjud bo'lsa, yangilash
                                for key, value in person_data.items():
                                    setattr(existing_person, key, value)
                                existing_person.save()
                            else:
                                # Yangi yaratish
                                existing_person = Supervisor.objects.create(**person_data)

                            nol_count = 7 - len(person_data['ps_num'])
                            ps_num = "0" * nol_count + person_data['ps_num']

                            data: dict = get_personal_data(imi=existing_person.imei, ps=f"{existing_person.ps_ser}{ps_num}")
                            if data['status'] == 1:
                                existing_person.img_b64 = data['photo']
                                existing_person.last_name = data['sname']
                                existing_person.first_name = data['fname']
                                existing_person.middle_name = data['mname']
                                existing_person.gender = 'M' if data['sex'] == 1 else 'F'
                                existing_person.save()
                            else:
                                messages.error(request, f"{data['message']}")
                            success_count += 1
                        except Exception as e:
                            error_count += 1
                            errors.append(f"Qator {index + 2}: {str(e)}")

                    # Xabar ko'rsatish
                    if success_count > 0:
                        messages.success(request, f"✅ Muvaffaqiyatli yuklandi: {success_count} ta shaxs")

                    if error_count > 0:
                        error_message = f"❌ Xatoliklar: {error_count} ta\n" + "\n".join(errors[:10])
                        if len(errors) > 10:
                            error_message += f"\n... va yana {len(errors) - 10} ta xatolik"
                        messages.error(request, error_message)

                    return redirect('..')

                except Exception as e:
                    messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'title': 'Excel fayldan shaxslarni yuklash',
            'site_header': 'Person Import',
            'has_view_permission': True,
        }
        return render(request, "admin/supervisor/import_excel.html", context)


    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Person Import"

        user = request.user

        headers = [
            'Familiya', 'Ism', 'Otasining ismi', 'Jinsi', 'Pasport seriya', 'Pasport raqam', 'PINFL',
        ]

        if user.is_admin or user.is_central:
            headers.append('Pcode')
            headers.append('Role')
        ws.append(headers)

        # Namuna ma'lumot
        sample_data = [
            'Oqboyev', 'Qoraboy', 'Sariq o\'g\'li', 'M', 'AA', '1234567', '12345678901234', '14', 'supevisor yoki staff'
        ]
        if user.is_delegate:
            sample_data.remove('14')
            sample_data.remove('supevisor yoki staff')
        ws.append(sample_data)

        # Sarlavhalarni qalin qilish
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Ustunlar kengligini sozlash
        column_widths = [15, 15, 20, 15, 8, 15, 15, 18, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Faylni saqlash
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=supervisor_template.xlsx'

        return response

    compressed_fields = True
    list_fullwidth = False
    list_filter_sheet = True
    list_horizontal_scrollbar_bottom = True
    paginator = InfinitePaginator
    show_full_result_count = True

    formfield_overrides = {
        models.ForeignKey: {
            "widget": UnfoldAdminSelectWidget,
        },
    }


@admin.register(EventSupervisor)
class EventSupervisorAdmin(ModelAdmin, ExportActionModelAdmin):
    list_display = ['id', 'supervisor', 'exam', 'zone', 'category_name', 'test_date', 'sm', 'group_n', 'is_participated']
    list_filter = ['is_participated', ]
    readonly_fields = ['id', 'created_at', 'updated_at', ]
    search_fields = ['supervisor__last_name', 'supervisor__first_name', 'supervisor__middle_name', 'supervisor__imei']
    list_display_links = ['id', 'supervisor', 'exam', 'zone', 'category_name']
    list_per_page = 30

    def get_queryset(self, request):
        qs = super().get_queryset(request).order_by('-id')
        user = User.objects.get(id=request.user.id)

        if user.is_admin or user.is_central:
            return qs

        user_region = getattr(request.user, "region", None)
        if user_region:
            return qs.filter(supervisor__region=user_region)
        return qs.none()

    # fieldsets = (
    #     ('Shaxsiy ma\'lumotlar', {
    #         'fields': ('supervisor__last_name', 'supervisor__first_name', 'supervisor__middle_name', 'supervisor__gender', 'supervisor__ps_ser', 'supervisor__ps_num', 'supervisor__imei',)
    #     }),
    #     ('Viloyat va bino', {
    #         'fields': ('supervisor__region', 'zone', )
    #     }),
    #     ('Holat', {
    #         'fields': ('supervisor__status', 'is_participated')
    #     }),
    #     ('Time', {
    #         'fields': ('created_at', 'updated_at')
    #     }),
    # )

    compressed_fields = True
    list_fullwidth = False
    list_filter_sheet = True
    list_horizontal_scrollbar_bottom = True
    paginator = InfinitePaginator
    show_full_result_count = True

    formfield_overrides = {
        models.ForeignKey: {
            "widget": UnfoldAdminSelectWidget,
        },
    }
